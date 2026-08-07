"""
Lecture Hall Roster Scheduling - Greedy Algorithm (Activity Selection)
------------------------------------------------------------------------
Problem: A university has ONE lecture hall. Several classes have
requested to use it on the same day, but some of their requested time
slots overlap. Given each class's requested start date/time and finish
date/time, select the MAXIMUM number of classes that can be scheduled
into the hall without any two classes overlapping.

Greedy Choice: At each step, always pick the class that finishes
EARLIEST among the classes still compatible with what has already
been scheduled. Finishing early leaves the most room in the roster
for other classes, which is why this greedy choice leads to an
optimal (maximum-count) solution.

INPUT: this program takes real input from the user, either:
  (1) a file - a CSV (.csv) or an Excel workbook (.xlsx), each row:
      ClassID, Course, Description, Date, Start, Finish
      Date must be DD-MM-YYYY and Start/Finish must be 24-hour HH:MM, or
  (2) typed directly on the screen, one class at a time.
No data is hardcoded into the source code.

Note: Quicksort is implemented manually to satisfy the requirement of
not using built-in sort functions.
"""

import re
import os
import datetime

TIME_PATTERN = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")           # 24-hour HH:MM
DATE_PATTERN = re.compile(r"^(0[1-9]|[12]\d|3[01])-(0[1-9]|1[0-2])-\d{4}$")  # DD-MM-YYYY


def _cell_to_str(value):
    """
    Converts a single Excel cell value to the plain string format this
    program expects. Excel often stores dates/times as real datetime
    objects (not text) even when the cell is displayed as DD-MM-YYYY
    or HH:MM, so those are reformatted here to match.
    """
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        # A full timestamp: could be a date-only or date+time cell.
        if value.time() == datetime.time(0, 0):
            return value.strftime("%d-%m-%Y")
        return value.strftime("%H:%M")
    if isinstance(value, datetime.date):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    return str(value).strip()


# ---------------------------------------------------------------------
# INPUT — Option 1: read requests from a file on disk (.csv or .xlsx)
# ---------------------------------------------------------------------
def read_requests_from_csv(filepath):
    """
    Reads class booking requests from a CSV file.
    Expected columns per line: ClassID,Course,Description,Date,Start,Finish
    The first line may optionally be a header ("ClassID,Course,...") and
    is skipped automatically if detected.
    """
    requests = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            if line_no == 1 and line.lower().startswith("classid"):
                continue  # skip header row
            fields = [x.strip() for x in line.split(",")]
            if len(fields) != 6:
                print(f"Skipping malformed line {line_no} in file (expected 6 fields): {line}")
                continue
            requests.append(fields)
    return requests


def read_requests_from_xlsx(filepath):
    """
    Reads class booking requests directly from an Excel (.xlsx) file
    using openpyxl. Expected columns (in order), one class per row:
    ClassID, Course, Description, Date, Start, Finish.
    The first row may optionally be a header and is skipped automatically.
    """
    from openpyxl import load_workbook

    requests = []
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None or all(c is None for c in row):
            continue
        fields = [_cell_to_str(c) for c in row[:6]]
        if row_no == 1 and fields[0].lower() == "classid":
            continue  # skip header row
        if len(fields) != 6:
            print(f"Skipping malformed row {row_no} in file (expected 6 columns).")
            continue
        requests.append(fields)
    return requests


def read_requests_from_file(filepath):
    """
    Dispatches to the correct reader based on the file extension
    (.xlsx -> Excel reader, anything else -> CSV reader).
    :param filepath: path to the file supplied by the user
    :return: list of raw request rows [id, course, description, date, start, finish]
    """
    if filepath.lower().endswith(".xlsx"):
        return read_requests_from_xlsx(filepath)
    return read_requests_from_csv(filepath)


# ---------------------------------------------------------------------
# INPUT — Option 2: read requests typed on the screen
# ---------------------------------------------------------------------
def get_valid_date(prompt):
    """Keeps asking until the user enters a valid DD-MM-YYYY date."""
    while True:
        d = input(prompt).strip()
        if DATE_PATTERN.match(d):
            return d
        print("Invalid date. Please use DD-MM-YYYY format, e.g. 03-08-2026")


def get_valid_time(prompt):
    """Keeps asking until the user enters a valid 24-hour HH:MM time."""
    while True:
        t = input(prompt).strip()
        if TIME_PATTERN.match(t):
            return t
        print("Invalid time. Please use 24-hour HH:MM format, e.g. 09:30")


def read_requests_from_screen():
    """
    Prompts the user to type in class booking requests one at a time.
    Press Enter on a blank Class ID to finish entering requests.
    :return: list of raw request rows [id, course, description, date, start, finish]
    """
    requests = []
    print("\nEnter class booking requests. Press Enter on a blank Class ID to stop.\n")
    while True:
        class_id = input("Class ID: ").strip()
        if not class_id:
            break
        course = input("Course Code: ").strip()
        description = input("Description: ").strip()
        date = get_valid_date("Date (DD-MM-YYYY): ")
        start = get_valid_time("Start Time (HH:MM): ")
        finish = get_valid_time("Finish Time (HH:MM): ")
        requests.append([class_id, course, description, date, start, finish])
        print(f"Added {class_id}.\n")
    return requests


def choose_input_source():
    """
    Asks the user whether to load requests from a file or type them on
    the screen, and returns the resulting list of raw requests.
    """
    while True:
        print("How would you like to enter the class booking requests?")
        print("  1. Load from a file")
        print("  2. Type them on the screen")
        choice = input("Enter 1 or 2: ").strip()
        if choice == "1":
            path = input("Enter the file path (e.g. roster_input_data.xlsx or .csv): ").strip()
            if not os.path.isfile(path):
                print(f"File not found: {path}\n")
                continue
            return read_requests_from_file(path)
        elif choice == "2":
            return read_requests_from_screen()
        else:
            print("Invalid choice, please enter 1 or 2.\n")


# ---------------------------------------------------------------------
# CONSTRAINTS / VALIDATION
# ---------------------------------------------------------------------
def validate_requests(requests, log=None):
    """
    Applies data-integrity constraints to the raw request list and
    splits it into (valid, invalid) lists.

    Constraints enforced:
      1. Class ID, Course Code, Description, Date must all be non-empty.
      2. Class ID must be unique across the whole roster.
      3. Date must match YYYY-MM-DD; Start and Finish must match 24-hour HH:MM.
      4. Start time must be strictly earlier than Finish time.

    :param requests: raw list of [id, course, description, date, start, finish]
    :param log: optional list to record [Class ID, Status, Reason]
    :return: (valid_requests, invalid_requests)
    """
    valid = []
    invalid = []
    seen_ids = set()

    for r in requests:
        class_id, course, description, date, start, finish = r
        reasons = []

        if not class_id.strip():
            reasons.append("Class ID is empty")
        if not course.strip():
            reasons.append("Course code is empty")
        if not description.strip():
            reasons.append("Description is empty")
        if not date.strip():
            reasons.append("Date is empty")
        elif not DATE_PATTERN.match(date):
            reasons.append(f"Date '{date}' is not valid DD-MM-YYYY")
        if class_id in seen_ids:
            reasons.append(f"Duplicate Class ID '{class_id}'")

        start_ok = bool(TIME_PATTERN.match(start))
        finish_ok = bool(TIME_PATTERN.match(finish))
        if not start_ok:
            reasons.append(f"Start time '{start}' is not valid 24-hour HH:MM")
        if not finish_ok:
            reasons.append(f"Finish time '{finish}' is not valid 24-hour HH:MM")
        if start_ok and finish_ok and start >= finish:
            reasons.append(f"Start time '{start}' is not before finish time '{finish}'")

        if reasons:
            invalid.append(r)
            if log is not None:
                log.append([class_id or "(blank)", "INVALID", "; ".join(reasons)])
        else:
            valid.append(r)
            seen_ids.add(class_id)
            if log is not None:
                log.append([class_id, "VALID", "Passed all constraints"])

    return valid, invalid


# ---------------------------------------------------------------------
# QUICKSORT (sorts requests by finish time, index 5)
# ---------------------------------------------------------------------
def partition(list_arr, low, high):
    """
    Method for quicksort algorithm
    :param list_arr: list of requests that needs to be sorted by finish time
    :param low: the lowest part of the list that this method will look at
    :param high: the highest part of the list that this method will look at
    :return: returns the position of the correctly sorted request in the list
    """
    pivot = list_arr[high][5]  # finish time of the last element is the pivot
    low_pos = low
    for i in range(low, high):
        if pivot > list_arr[i][5]:
            list_arr[low_pos], list_arr[i] = list_arr[i], list_arr[low_pos]
            low_pos += 1
    list_arr[high], list_arr[low_pos] = list_arr[low_pos], list_arr[high]
    return low_pos


def quick_sort(list_arr, low, high):
    """
    Quicksort algorithm to sort requests by finish time
    :param list_arr: the list of requests that will be sorted
    :param low: the lower boundary of the part of the list that needs to be sorted
    :param high: the upper boundary of the part of the list that needs to be sorted
    """
    if low < high:
        pivot_pos = partition(list_arr, low, high)
        quick_sort(list_arr, low, pivot_pos - 1)
        quick_sort(list_arr, pivot_pos + 1, high)


# ---------------------------------------------------------------------
# GREEDY ACTIVITY SELECTION
# ---------------------------------------------------------------------
def select_roster(sorted_requests, log=None):
    """
    Applies the greedy algorithm to select the maximum number of
    non-overlapping class bookings from an ALREADY SORTED, ALREADY
    VALIDATED list of requests.
    """
    if not sorted_requests:
        return []

    selected = [sorted_requests[0]]
    last_finish_time = sorted_requests[0][5]

    if log is not None:
        log.append([sorted_requests[0][0], "ACCEPTED", "Hall is free"])

    for i in range(1, len(sorted_requests)):
        current = sorted_requests[i]
        start_time = current[4]

        if start_time >= last_finish_time:
            selected.append(current)
            if log is not None:
                log.append([current[0], "ACCEPTED", f"Starts at/after {last_finish_time}"])
            last_finish_time = current[5]
        else:
            if log is not None:
                log.append([current[0], "REJECTED", f"Clashes until {last_finish_time}"])

    return selected


def print_table(title, requests):
    print(f"\n{title}")
    header = (f"{'Class':<8}| {'Course':<10}| {'Description':<38}"
              f"| {'Date':<12}| {'Start':<8}| {'Finish':<8}")
    print(header)
    print("-" * len(header))
    for r in requests:
        print(f"{r[0]:<8}| {r[1]:<10}| {r[2]:<38}| {r[3]:<12}| {r[4]:<8}| {r[5]:<8}")


def print_decision_log(title, log, col2="Decision"):
    print(f"\n{title}")
    header = f"{'Class':<8}| {col2:<10}| {'Reason':<55}"
    print(header)
    print("-" * len(header))
    for class_id, decision, reason in log:
        print(f"{class_id:<8}| {decision:<10}| {reason:<55}")


def main():
    print("=" * 60)
    print("    LECTURE HALL ROSTER SCHEDULING (GREEDY - ACTIVITY SELECTION)")
    print("=" * 60)

    print("\nProblem:")
    print("  One lecture hall is available. Several classes have each")
    print("  requested a time slot, but some of those slots overlap -")
    print("  the hall can only host one class at a time.")
    print("  Goal: schedule the MAXIMUM number of non-overlapping classes.")
    print()

    raw_requests = choose_input_source()

    if not raw_requests:
        print("\nNo requests entered. Nothing to schedule.")
        return

    print_table("\nAll Requested Bookings (raw input):", raw_requests)

    print("\nStep 0 - Validate constraints:")
    print("  Every request must have all fields filled in, a unique Class ID,")
    print("  a valid DD-MM-YYYY date, properly formatted 24-hour HH:MM times,")
    print("  and Start < Finish. Anything that fails is rejected here and")
    print("  never reaches scheduling.")
    validation_log = []
    valid_requests, invalid_requests = validate_requests(raw_requests, log=validation_log)
    print_decision_log("Validation Results:", validation_log, col2="Status")

    if invalid_requests:
        print(f"\n{len(invalid_requests)} request(s) rejected at validation and excluded from scheduling.")

    if not valid_requests:
        print("\nNo valid requests remain after validation. Nothing to schedule.")
        return

    print("\nStep 1 - Sort by finish time (Quicksort):")
    print("  Every VALID request is sorted so the class that finishes EARLIEST")
    print("  comes first. Considering early finishers first leaves the")
    print("  most remaining time in the day for other classes.")
    sorted_requests = [r[:] for r in valid_requests]  # copy so original list is unmutated
    quick_sort(sorted_requests, 0, len(sorted_requests) - 1)
    print_table("Valid Requests Sorted by Finish Time:", sorted_requests)

    print("\nStep 2 - Greedily accept or reject each class in that order:")
    print("  A class is ACCEPTED only if its start time is not earlier")
    print("  than the finish time of the last class already booked into")
    print("  the hall (i.e. it does not clash with the current booking).")
    decision_log = []
    selected = select_roster(sorted_requests, log=decision_log)
    print_decision_log("Accept / Reject Decisions:", decision_log)

    print_table("Final Lecture Hall Roster (Optimal Schedule):", selected)

    print(f"\nMaximum number of classes that can use the hall: {len(selected)}")
    print("Roster order:", " -> ".join(r[0] for r in selected))
    print("\nWhy this is optimal: this is the classic Activity Selection")
    print("problem. Always picking the next class with the earliest")
    print("finish time (and skipping anything that clashes) is proven to")
    print("never do worse than any other selection order, so the count")
    print("of classes scheduled above is the maximum possible.")


if __name__ == "__main__":
    main()
